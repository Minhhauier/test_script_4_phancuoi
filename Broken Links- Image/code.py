import time

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import requests


def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def ktra_hinhanh():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch()
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.wait_for_load_state("load")
        page.locator("text=Elements").click()
        page.locator("text=Broken Links - Images").click()
        images = page.locator("img").element_handles()  # chọn tất cả các image trong trang hiện tại
        desired_images = images[2:4]
        for index, img in enumerate(desired_images):
            # Lấy src của hình ảnh bằng JavaScript
            img_src = img.evaluate("el => el.src")
            print(f"Checking image {index + 1}: {img_src}")
            response = requests.get(img_src)
            # Kiểm tra nếu src trống
            if response.headers['Content-Type'].startswith('image'):
                print(f"PASS: Image {index + 1} loaded successfully.")
                write_to_excel("test_results.xlsx", [
                    f"Test image 1", "Pass", f"Image {index + 1} loaded successfully"
                ])
            else:
                print(f"FAIL: Image {index + 1} has an empty or missing src.")
                write_to_excel("test_results.xlsx", [
                    f"Test image 1", "Fail", f"Image {index + 1} has an empty or missing src."
                ])
        browser.close()
def ktra_link():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.wait_for_load_state("load")
        page.locator("text=Elements").click()
        page.locator("text=Broken Links - Images").click()
        links=[{"locator" : "Click Here for Valid Link"},
               {"locator" : "Click Here for Broken Link"}]
        for link in links:
            try:
                page_new = page.locator(f"text={link['locator']}")
                url = page_new.get_attribute('href')
                response = requests.get(url)
                if response.status_code == 200 and link['locator'] == "Click Here for Valid Link":
                    print(f"Pass: {response} valid link hoạt dộng chính xác")
                    write_to_excel("test_results.xlsx", [
                        "Test Valid link","Pass",f"{response} valid link hoạt dộng chính xác"
                    ])
                elif response.status_code !=200 and link['locator'] == "Click Here for Broken Link":
                    print(f"Pass {response} broken link hoạt động chính xác")
                    write_to_excel("test_results.xlsx", [
                        "Test broken link","Pass",f"{response} broken link hoạt dộng chính xác"
                    ])
                else:
                    print(f"Fail {response} không họat động chính xác")
                    write_to_excel("test_results.xlsx", [
                        "Test link","Fail",f"{response} link hoạt động không chính xác"
                    ])
            except Exception as e:
                print(f"Error: Không xác định được link cần kiểm tra- {e}")
        browser.close()


def main():
    ktra_link()
    time.sleep(1)
    ktra_hinhanh()


if __name__ == "__main__":
    main()

