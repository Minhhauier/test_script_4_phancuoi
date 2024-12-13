import time

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
from pathlib import Path

def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def ktra_upload():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch()
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.locator("text=Elements").click()
        page.locator("text=Upload and Download").click()
        page.set_input_files("#uploadFile", "C:/demokthu/Sheet.xlsx")
        uploaded_file_path = page.locator("#uploadedFilePath").inner_text()
        if os.path.basename("C:/demokthu/Sheet.xlsx") in uploaded_file_path:
            print("PASS: Tệp được tải lên thành công và hiển thị đúng tên.")
            write_to_excel("test_results.xlsx", [
                "Test Upload", "Pass", f"Tệp được tải lên thành công và hiển thị đúng tên"
            ])
        else:
            print("FAIL: Tệp không được hiển thị đúng sau khi tải lên.")
            write_to_excel("test_results.xlsx", [
                "Test Upload", "Fail", f"Tệp được tải lên không thành công hoặc không hiển thị đúng tên"
            ])

        browser.close()

def ktra_download():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.wait_for_load_state("load")
        page.locator("text=Elements").click()
        page.locator("text=Upload and Download").click()
        current_dir = Path(__file__).parent
        new_page = page.locator("#downloadButton")
        with page.expect_download() as download_info:
            new_page.click()
        value = download_info.value
        download_path = current_dir / value.suggested_filename
        value.save_as(str(download_path))
        try:
            assert download_path.exists(), "File download không thành công"
            with open(download_path, "rb") as file:
                content = file.read()
                expected_size = 4096  # Số byte mong đợi
                assert len(content) == expected_size, f"kích thước file không đúng. Expected {expected_size}, got {len(content)}"
            print(f'Pass: File download thành công')
            write_to_excel("test_results.xlsx", [
                "Test Download file", "Pass", f" File Download thành công"
            ])
        except Exception as e:
            print(f'Fail: {e}')
            write_to_excel("test_results.xlsx", [
                "Test Download file", "Fail", f" File Download không thành công"
            ])
        browser.close()


def main():
    ktra_download()
    time.sleep(1)
    ktra_upload()


if __name__ == "__main__":
    main()