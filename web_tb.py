from playwright.sync_api import sync_playwright

import time
from openpyxl import Workbook, load_workbook
import os
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def test_add_data_to_table():
    with sync_playwright() as p:
        # Khởi tạo trình duyệt
        browser = p.chromium.launch(headless=False)  # Nếu bạn muốn xem giao diện kiểm thử
        page = browser.new_page()

        # Truy cập trang DemoQA
        page.goto('https://demoqa.com/webtables')
        page.wait_for_load_state("load")
        # Nhấn vào nút "Add"
        page.locator('#addNewRecordButton').click()

        # Nhập các giá trị
        page.locator("#firstName").fill('Nguyễn')
        page.locator("#lastName").fill('Bình')
        page.locator("#userEmail").fill('nguyenbinh@example.com')
        page.locator("#age").fill('30')
        page.locator("#salary").fill('1000')
        page.locator("#department").fill('IT')

        # Nhấn nút Submit

        page.locator('text=Submit').click()

        # Lặp qua từng dòng
        rows = page.locator('div.rt-tr-group')
        # Lặp qua từng dòng
        found = False
        for i in range(rows.count()):
            row = rows.nth(i)

            # Lấy toàn bộ nội dung dòng và làm sạch
            row_text = row.all_inner_texts()[0].replace('\n', ' ').strip()
            print(f"Dòng {i + 1}: {row_text}")

            # Kiểm tra nếu dòng chứa tất cả các giá trị mong muốn
            if (
                    "Nguyễn Bình" in row_text and
                    "30" in row_text and
                    "nguyenbinh@example.com" in row_text and
                    "1000" in row_text and
                    "IT" in row_text
            ):
                found = True
                break

        assert found, "Không tìm thấy dòng chứa dữ liệu mong muốn."
        write_to_excel("test_results.xlsx", [
            "Test thêm dữ liệu vào web Tables", "Pass", f"Thêm thành công dữ liệu vào bảng"
        ])
        print("Pass: Thêm thành công dữ liệu vào bảng")
        # Đóng trình duyệt
        browser.close()

def test_edit_data():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # Mở trình duyệt trong chế độ hiển thị
        page = browser.new_page()

        # Truy cập trang web
        page.goto("https://demoqa.com/webtables")
        rows = page.locator('div.rt-tr-group')
        row = rows.nth(0)
        row_text = row.all_inner_texts()[0].replace('\n', ' ').strip()
        print("Thong tin dong 1 truoc khi thay doi")
        print(f"Dòng 1: {row_text}")
        # Nhấn vào nút "Edit" của dòng đầu tiên
        edit_button = page.locator('#edit-record-1')
        assert edit_button.is_visible(), "Nút Edit không hiển thị."
        edit_button.click()

        # Điền thông tin vào các trường
        page.fill("#firstName", "Nguyễn")  # First Name
        page.fill("#lastName", "Hoàng")   # Last Name
        page.fill("#age", "35")           # Age
        page.fill("#userEmail", "nguyenhoang@example.com")  # Email
        page.fill("#salary", "1500")      # Salary
        page.fill("#department", "HR")    # Department

        # Nhấn nút Submit để lưu thay đổi
        submit_button = page.locator('button#submit')
        assert submit_button.is_visible(), "Nút Submit không hiển thị."
        submit_button.click()

        # Kiểm tra dữ liệu đã được thay đổi trong bảng

        found = False
        for i in range(rows.count()):
            row = rows.nth(i)
            row_text = row.all_inner_texts()[0].replace('\n', ' ').strip()
            print("Thong tin dong 1 sau khi thay doi")
            print(f"Dòng {i+1}: {row_text}")

            if (
                "Nguyễn" in row_text and
                "Hoàng" in row_text and
                "35" in row_text and
                "nguyenhoang@example.com" in row_text and
                "1500" in row_text and
                "HR" in row_text
            ):
                found = True
                break

        assert found, "Không tìm thấy dữ liệu đã chỉnh sửa trong bảng."
        print("Pass: Thay đổi thông tin thành công!")
        write_to_excel("test_results.xlsx", [
            "Test thay đổi thông tin trong tables", "Pass", f"Thay đổi thông tin thành công"
        ])

        browser.close()
if __name__ == '__main__':
    test_add_data_to_table()
    time.sleep(1)
    test_edit_data()
