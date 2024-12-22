from playwright.sync_api import sync_playwright
import time
from openpyxl import Workbook, load_workbook
import os


# Hàm ghi kết quả vào file Excel
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa, nếu chưa thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # Nếu đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()


# Hàm kiểm thử radio button
def test_radio_button_selection():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/radio-button")

        # Bước 1: Chọn radio button Yes và kiểm tra thông báo
        page.click('label[for="yesRadio"]')
        time.sleep(1)
        ketqua = page.text_content('.text-success')

        try:
            # Kiểm tra thông báo đã hiển thị sau khi chọn Yes
            assert ketqua == "You have selected Yes", "Thông báo không đúng cho Yes"
            print("Thông báo hiển thị đúng khi chọn 'Yes'.")
            write_to_excel("test_results.xlsx", ["Test Select 'Yes'", "Pass", "Thông báo hiển thị đúng khi chọn 'Yes'."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'Yes'", "Fail", str(e)])

        # Bước 2: Chọn radio button Impressive và kiểm tra thông báo
        page.click('label[for="impressiveRadio"]')
        time.sleep(1)
        ketqua = page.text_content('.text-success')

        try:
            # Kiểm tra thông báo đã hiển thị sau khi chọn Impressive
            assert ketqua == "You have selected Impressive", "Thông báo không đúng cho Impressive"
            print("Thông báo hiển thị đúng khi chọn 'Impressive'.")
            write_to_excel("test_results.xlsx", ["Test Select 'Impressive'", "Pass", "Thông báo hiển thị đúng khi chọn 'Impressive'."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'Impressive'", "Fail", str(e)])

        # Bước 3: Kiểm tra rằng radio button No không thể chọn
        is_no_enabled = page.is_enabled('label[for="noRadio"]')
        try:
            assert not is_no_enabled, "Radio button No should not be selectable"
            print("Radio button 'No' không thể chọn.")
            write_to_excel("test_results.xlsx", ["Test Select 'No'", "Pass", "Radio button 'No' không thể chọn."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test Select 'No'", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else:  # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()

def test_radio_buttons():
    with sync_playwright() as p:
        # Mở trình duyệt Chrome
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # Truy cập trang
        page.goto("https://demoqa.com/radio-button")

        # Kiểm tra radio button "Yes"
        try:
            page.click('label[for="yesRadio"]')
            assert "You have selected Yes" in page.inner_text('#app'), "Thông báo không đúng cho Yes"
            print("Thông báo hiển thị đúng khi chọn 'Yes'.")
            write_to_excel("test_results.xlsx", ["Test radio button 'Yes'", "Pass", "Thông báo hiển thị đúng khi chọn 'Yes'"])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'Yes'", "Fail", str(e)])

        # Kiểm tra radio button "Impressive"
        try:
            page.click('label[for="impressiveRadio"]')
            assert "You have selected Impressive" in page.inner_text('#app'), "Thông báo không đúng cho Impressive"
            print("Thông báo hiển thị đúng khi chọn 'Impressive'.")
            write_to_excel("test_results.xlsx", ["Test radio button 'Impressive'", "Pass", "Thông báo hiển thị đúng khi chọn 'Impressive'"])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'Impressive'", "Fail", str(e)])

        # Kiểm tra rằng "No" không thể chọn
        try:
            no_button = page.locator('label[for="noRadio"]')
            assert no_button.is_disabled(), "Radio button 'No' không bị vô hiệu hóa"
            print("'No' không thể chọn vì bị vô hiệu hóa.")
            write_to_excel("test_results.xlsx", ["Test radio button 'No'", "Pass", "'No' không thể chọn vì bị vô hiệu hóa."])
        except AssertionError as e:
            print(f"Test failed: {str(e)}")
            write_to_excel("test_results.xlsx", ["Test radio button 'No'", "Fail", str(e)])

        # Đóng trình duyệt
        browser.close()
# Gọi hàm kiểm thử
test_radio_buttons()