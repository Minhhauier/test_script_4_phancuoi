from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import requests
import time
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def test_checkbox_page():
    with sync_playwright() as p:
        # Mở trình duyệt
        browser = p.chromium.launch()  # Đặt headless=True nếu muốn chạy ẩn
        context = browser.new_context()
        page = context.new_page()
        url = "https://demoqa.com/checkbox"
        page.goto(url)
        try:
            assert page.title() == "ToolsQA"
            print("✅ Trang được tải chính xác.")
        except AssertionError:
            print("❌ Trang không tải đúng hoặc title không khớp.")
            write_to_excel("test_results.xlsx", [
                f"Test checkbox", "Fail", f"Image has an empty or missing src."
            ])
        checkbox_selector = ".rct-checkbox"
        checkboxes = page.query_selector_all(checkbox_selector)
        if len(checkboxes) > 0:
            print(f"✅ Có {len(checkboxes)} checkbox được hiển thị.")
        else:
            print("❌ Không tìm thấy checkbox nào trên trang.")
        errors = page.query_selector_all(".error, .alert")
        if len(errors) > 0:
            print("❌ Trang có lỗi giao diện hoặc thông báo bất thường.")
        else:
            print("✅ Không phát hiện lỗi giao diện hoặc thông báo bất thường.")
        browser.close()
def testcase_1():
    with sync_playwright() as p:
        # Mở trình duyệt
        browser = p.chromium.launch()  # headless=False để thấy giao diện
        context = browser.new_context()
        page = context.new_page()

        # Bước 1: Mở URL
        url = "https://demoqa.com/text-box"
        page.goto(url)

        # Bước 2: Kiểm tra trang được tải đầy đủ
        try:
            assert page.title() == "ToolsQA"
            print("✅ Trang được tải đầy đủ và tiêu đề chính xác.")
        except AssertionError:
            print("❌ Tiêu đề trang không đúng hoặc trang không tải đầy đủ.")
            write_to_excel("test_results.xlsx", [
                f"test UI textbox", "Fail", f"❌ Tiêu đề trang không đúng hoặc trang không tải đầy đủ"
            ])
        # Bước 3: Kiểm tra giao diện và các thành phần
        try:
            # Kiểm tra trường nhập liệu
            input_fields = ["#userName", "#userEmail", "#currentAddress", "#permanentAddress"]
            for field in input_fields:
                assert page.query_selector(field) is not None
            print("✅ Các trường nhập liệu hiển thị đầy đủ.")

            # Kiểm tra nút bấm "Submit"
            assert page.query_selector("#submit") is not None
            print("✅ Nút bấm 'Submit' hiển thị đúng vị trí.")
        except AssertionError as e:
            print("❌ Giao diện không hiển thị đầy đủ hoặc các thành phần bị thiếu.")
            write_to_excel("test_results.xlsx", [
                f"test UI textbox", "Fail", f"❌ Giao diện không hiển thị đầy đủ hoặc các thành phần bị thiếu"
            ])

        # Kiểm tra lỗi giao diện (nếu có)
        errors = page.query_selector_all(".error, .alert")
        if len(errors) > 0:
            print("❌ Trang có lỗi hiển thị hoặc thông báo bất thường.")
        else:
            print("✅ Không phát hiện lỗi giao diện hoặc thông báo bất thường.")
        write_to_excel("test_results.xlsx", [
            f"test UI textbox", "Pass", f"✅ Không phát hiện lỗi giao diện hoặc thông báo bất thường"
        ])
        # Đóng trình duyệt
        browser.close()

def testcase_2():
    with sync_playwright() as p:
        # Mở trình duyệt
        browser = p.chromium.launch()  # headless=False để quan sát
        context = browser.new_context()
        page = context.new_page()

        # Bước 1: Truy cập URL
        url = "https://demoqa.com/text-box"
        page.goto(url)

        # Bước 2: Nhập dữ liệu vào các trường
        test_data = {
            "Full Name": "Nguyễn Văn A",
            "Email": "nguyen.a@example.com",
            "Current Address": "123 Đường ABC, Quận 1, TP HCM",
            "Permanent Address": "456 Đường DEF, Quận 2, TP HCM"
        }

        # Nhập dữ liệu vào từng trường
        page.fill("#userName", test_data["Full Name"])
        page.fill("#userEmail", test_data["Email"])
        page.fill("#currentAddress", test_data["Current Address"])
        page.fill("#permanentAddress", test_data["Permanent Address"])

        # Bước 3: Kiểm tra dữ liệu đã nhập
        try:
            assert page.input_value("#userName") == test_data["Full Name"]
            assert page.input_value("#userEmail") == test_data["Email"]
            assert page.input_value("#currentAddress") == test_data["Current Address"]
            assert page.input_value("#permanentAddress") == test_data["Permanent Address"]
            print("✅ Dữ liệu nhập chính xác vào tất cả các trường.")
            write_to_excel("test_results.xlsx", [
                f"test nhap du lieu vao textbox", "Pass", f"✅ Dữ liệu nhập chính xác vào tất cả các trường. "
            ])
        except AssertionError:
            print("❌ Dữ liệu không khớp với thông tin đã nhập.")
            write_to_excel("test_results.xlsx", [
                f"test nhap du lieu vao textbox", "Fail", f"❌ Dữ liệu không khớp với thông tin đã nhập. "
            ])

        # Đóng trình duyệt
        browser.close()
def testcase_4():
    with sync_playwright() as p:
        # Mở trình duyệt
        browser = p.chromium.launch()  # headless=False để quan sát
        context = browser.new_context()
        page = context.new_page()

        # Bước 1: Truy cập URL
        url = "https://demoqa.com/text-box"
        page.goto(url)

        # Bước 2: Nhập dữ liệu vào các trường
        test_data = {
            "Full Name": "Nguyễn Văn A",
            "Email": "nguyen.a@example.com",
            "Current Address": "123 Đường ABC, Quận 1, TP HCM",
            "Permanent Address": "456 Đường DEF, Quận 2, TP HCM"
        }

        page.fill("#userName", test_data["Full Name"])
        page.fill("#userEmail", test_data["Email"])
        page.fill("#currentAddress", test_data["Current Address"])
        page.fill("#permanentAddress", test_data["Permanent Address"])

        # Bước 3: Nhấn nút Submit
        page.click("#submit")

        # Bước 4: Kiểm tra bảng kết quả hiển thị đúng thông tin
        try:
            result_full_name = page.text_content("#name")
            result_email = page.text_content("#email")
            result_current_address = page.text_content("p#currentAddress")
            result_permanent_address = page.text_content("p#permanentAddress")

            assert "Name:" + test_data["Full Name"] in result_full_name
            assert "Email:" + test_data["Email"] in result_email
            assert test_data["Current Address"] in result_current_address
            assert test_data["Permanent Address"] in result_permanent_address

            print("✅ Nút Submit hoạt động đúng, bảng kết quả hiển thị chính xác.")
            write_to_excel("test_results.xlsx", [
                "Test nut submit cua  textbox", "Pass", "✅ Nút Submit hoạt động đúng, bảng kết quả hiển thị chính xác."])
        except AssertionError:
            print("❌ Thông tin trong bảng kết quả không khớp với dữ liệu đã nhập.")
            write_to_excel("test_results.xlsx", [
                "Test nut submit cua  textbox", "Fail", "❌ Thông tin trong bảng kết quả không khớp với dữ liệu đã nhập."])
        # Đóng trình duyệt
        browser.close()


if __name__ == "__main__":
    testcase_1()
    time.sleep(1)
    testcase_2()
    time.sleep(1)
    testcase_4()