import time

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def ktra_home():
    with sync_playwright() as playwright:
        try:
            browser = playwright.chromium.launch(headless=False)
            context = browser.new_context()
            page = context.new_page()
            page.goto("https://demoqa.com/links",timeout=60000)
            Home_page = page.locator("#simpleLink")
            with page.expect_popup() as new_page:
                Home_page.click()
            page_test = new_page.value
            page_test.wait_for_load_state("load")
            assert page_test.url.rstrip("/") == "https://demoqa.com", f"False"
            print("Pass")
            browser.close()
            #ghi kết quả kiểm tra đúng vào Excel
            write_to_excel("test_results.xlsx", ["Test Click Home Link", "Pass", "Tab mới mở đúng URL!"])
        except AssertionError as e:
            # Ghi kết quả kiểm tra thất bại vào Excel
            write_to_excel("test_results.xlsx", ["Test Click Home Link", "Fail", str(e)])
        finally:
            browser.close()

def ktra_api():
    thunghiem = " "
    with sync_playwright() as playwriht:
        browser = playwriht.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto("https://demoqa.com/links")
        links = [
                {"selector": "#created", "expected_status": "201"},
                {"selector": "#no-content", "expected_status": "204"},
                {"selector": "#moved", "expected_status": "301"},
                {"selector": "#bad-request", "expected_status": "400"},
                {"selector": "#unauthorized", "expected_status": "401"},
                {"selector": "#forbidden", "expected_status": "403"},
                {"selector": "#invalid-url", "expected_status": "404"}
            ]
        for link in links:
            new = page.locator(link['selector'])
            new.click()
            try:
                dem=10
                phanhoi = page.locator("#linkResponse").text_content(timeout=5000)
                while phanhoi == thunghiem and dem > 0:
                    phanhoi = page.locator("#linkResponse").text_content(timeout=5000)
                    time.sleep(0.1)
                    dem=dem-1
                if link['expected_status'] in phanhoi:
                    print(f"PASS: {link['selector']} returned expected status.")
                    write_to_excel("test_results.xlsx", [
                        f"Test {link['selector']}", "Pass", f"Hiển thị đúng {phanhoi}"
                    ])
                else:
                    print(f"FAIL: {link['selector']} did not return expected status.")
                    write_to_excel("test_results.xlsx", [
                        f"Test {link['selector']}", "Failed", f"Expected {link['expected_status']} but got {phanhoi}"
                    ])
                thunghiem = phanhoi
            except Exception as e:
                print(f"ERROR: Unable to verify {link['selector']} - {e}")
                write_to_excel("test_results.xlsx", [
                    f"Test {link['selector']}", "Error", f"Exception: {e}"
                ])

        browser.close()


def main():
    ktra_home()
    time.sleep(1)
    ktra_api()


if __name__ == "__main__":
    main()


