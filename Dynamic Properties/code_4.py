import time

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
def write_to_excel(file_name, row_data):
    # Kiểm tra xem file Excel đã tồn tại chưa nếu chưa tồn tại thì tạo mới và thêm tiêu đề các cột
    if not os.path.exists(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Test Name", "Result", "Details"])
    else: # nêú đã tồn tại thì mở ra và cho hoạt động
        workbook = load_workbook(file_name)
        sheet = workbook.active

    # Thêm dữ liệu kiểm tra ra một hàng mới
    sheet.append(row_data)
    workbook.save(file_name)
    workbook.close()
def kichban1():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.locator("text=Elements").click()
        page.locator('text=Dynamic Properties').click()
        is_disabled = page.locator('#enableAfter').is_disabled()
        try:
            assert is_disabled, "Nút không bị vô hiệu hóa khi tải trang!"
            time.sleep(5)
            is_enabled = page.locator('#enableAfter').is_enabled()
            assert is_enabled, "Nút không được kích hoạt sau 5 giây!"
            print(f'Pass: Nút bị vô hiệu hóa khi tải trang và được kích hoạt sau 5 giây')
            write_to_excel("test_results.xlsx", [
                "Test trạng thái của nút", "Pass", "Nút bị vô hiệu hóa khi tải trang và được kích hoạt sau 5 giây"
            ])
        except Exception as e:
            print(f'Fail {e}')
            write_to_excel("test_results.xlsx", [
                "Test trạng thái của nút", "Fail", f"{e}"
            ])
        browser.close()
def kichban2():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.locator("text=Elements").click()
        page.locator('text=Dynamic Properties').click()
        color = page.locator('#colorChange').evaluate("element => getComputedStyle(element).color")
        time.sleep(5)
        new_color = page.locator('#colorChange').evaluate("element => getComputedStyle(element).color")
        if color == new_color:
            print("Fail: Nút không đổi màu sau 5 giây!")
            write_to_excel("test_results.xlsx", [
                "Test khả năng đổi màu", 'fail', "Nút không đổi màu sau 5 giây !"
            ])
        else:
            print("Pass: Nút đổi màu sau 5 giây!")
            write_to_excel("test_results.xlsx", [
                "Test khả năng đổi màu", "Pass", "Nút đổi màu sau 5 giây !"
            ])
        browser.close()

def kichban3():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        page.goto('https://demoqa.com/')
        page.locator("text=Elements").click()
        page.locator('text=Dynamic Properties').click()
        button = page.locator('#visibleAfter').is_visible()
        try:
            assert not button, "Nút hiển thị từ đầu !"
            time.sleep(5)
            button = page.locator('#visibleAfter').is_visible()
            assert button, "Nút không hiển thị sau 5 giây !"
            print(f'Pass: Nút xuất hiện sau 5 giây !')
            write_to_excel("test_results.xlsx", [
                "Test hiển thị phần tử tự động", "Pass", "Nút hiển thị sau 5 giây"
            ])
        except Exception as e:
            print(f'Fail {e}')
            write_to_excel("test_results.xlsx", [
                "Test hiển thị phần tử tự động ", "Fail", f"{e}"
            ])
        browser.close()


def main():
    kichban1()
    time.sleep(1)
    kichban2()
    time.sleep(1)
    kichban3()


if __name__=="__main__":
    main()




